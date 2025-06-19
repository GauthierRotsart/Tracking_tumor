import numpy as np
import torch
import hydra
import random
import pandas as pd
import torch.nn as nn

from tqdm.auto import tqdm
from openpyxl import load_workbook
from refactor_TF import Net
from refactor_functional import load_data_structure, data_array_creation, denormalisation, to_device, get_default_device, \
    subsequent_mask, association, association_image
 
 
def inference_random(model, cfg, test_part, verbose=False):
    base_path = '/Benson_DATA1/grotsart/'
    translate_path = '/linux/grotsartdehe/translate.xlsx'
    amplitude_path = '/linux/grotsartdehe/amplitude.xlsx'
    if cfg.multi is False:
        saving_path = f'/linux/grotsartdehe/Medical_physics/{cfg.training.params.width}-{cfg.training.params.patch_size}/' \
                      f'LR-{cfg.training.params.learning_rate}/PS-scratch_{cfg.data_quantity}_epoch_' \
                      f'{cfg.inference.params.epoch}.xlsx'
    else:
        saving_path = f'/linux/grotsartdehe/Medical_physics/{cfg.training.params.width}-{cfg.training.params.patch_size}/' \
                      f'LR-{cfg.training.params.learning_rate}/MP-scratch_{cfg.data_quantity}_epoch_' \
                      f'{cfg.inference.params.epoch}.xlsx'
    if cfg.patient < 16:
        patient = cfg.patient
        root_test = f'{base_path}FDGorFAZA_study/Patient_{patient}/{list(test_part)[-1]}/FDG{list(test_part)[-1]}/{cfg.training.params.regularity}/test_images_{list(test_part)[-1]}1/resized/{cfg.training.params.width}/normalised/'
        translate_data = pd.read_excel(translate_path, header=None, sheet_name='FDG1' if test_part == 'T1' else 'FDG1').values
        amplitude_data = pd.read_excel(amplitude_path, header=None, sheet_name='FDG1').values
    else:
        patient = cfg.patient - 15
        root_test = f'{base_path}CPAP_study/Patient_{patient}/{list(test_part)[-1]}/NO_CPAP/' \
                    f'{cfg.training.params.regularity}/test_images_{list(test_part)[-1]}1/resized/' \
                    f'{cfg.training.params.width}/normalised/'
        translate_data = pd.read_excel(translate_path, header=None, sheet_name='NO_CPAP1' if test_part == 'T1' else 'NO_CPAP2_reg').values
        amplitude_data = pd.read_excel(amplitude_path, header=None, sheet_name='NO_CPAP1').values
 
    translation = [translate_data[patient, 1], translate_data[patient, 2], translate_data[patient, 3]]
    amplitude = [amplitude_data[patient, 1], amplitude_data[patient, 2], amplitude_data[patient, 3]]
    batch_size = cfg.training.params.batch_size
    width = cfg.training.params.width
    height = cfg.training.params.height
    patch_size = cfg.training.params.patch_size
    patches_sequence = (height // patch_size) * (width // patch_size)
 
    loss_function = nn.MSELoss()
    model.eval()
    with torch.no_grad():
        shuffle_batch = True
 
        with tqdm(range(10), total=10, unit="serie", desc=f"Test loop") as pbar:
            nbr_batch = 0
            epoch_loss = []
            epoch_loss_x = []
            epoch_loss_y = []
            epoch_loss_z = []
            for idxSerie in range(10):
                selected_index = [i for i in range(0, 100)]
                if shuffle_batch:
                    random.shuffle(selected_index)
 
                data_path = f'{root_test}serie{idxSerie}/Patient_{patient}_1500_DRRMasksAndCOM_serie_{idxSerie}.p'
                dynamic_sequence = load_data_structure(data_path)
                x_set, y_set = data_array_creation(dynamic_sequence, cfg, translation, amplitude)
                for idx in range(0, len(selected_index), batch_size):
                    batch_idx = selected_index[idx:idx + batch_size]
                    batch_x = np.zeros(
                        (len(batch_idx), cfg.training.params.img_sequence, patches_sequence,
                         cfg.training.params.patch_size ** 2))
                    batch_y = np.zeros(
                        (len(batch_idx), cfg.training.params.horizon, cfg.training.params.output_dimension))
                    b_i = 0
                    for i in batch_idx:
                        x = association_image(x_set, i, cfg.training.params.img_sequence)
                        y = association(y_set, i + cfg.training.params.img_sequence, cfg.training.params.horizon)
                        batch_x[b_i, :, :, :] = x
                        batch_y[b_i, :, :] = y
                        b_i += 1
                    enc_inp = torch.from_numpy(batch_x).float().to(cfg.device)
                    target = torch.from_numpy(batch_y).float().to(cfg.device)
                    start_of_sequence = torch.Tensor([0, 0, 0, 1]).unsqueeze(0).unsqueeze(1).to(cfg.device)
                    start_of_sequence = start_of_sequence.repeat(len(batch_idx), 1, 1)

                    target_c = torch.zeros((target.shape[0], target.shape[1], 1)).to(cfg.device)
                    target = torch.cat((target, target_c), -1)

                    dec_inp = start_of_sequence
                    for i in range(cfg.training.params.horizon):
                        enc_mask = torch.ones((enc_inp.shape[0], 1, enc_inp.shape[1] * enc_inp.shape[2])).to(cfg.device)
                        tgt_mask = subsequent_mask(dec_inp.shape[1]).repeat(dec_inp.shape[0], 1, 1).to(cfg.device)
                        pred = model(enc_inp, enc_mask, dec_inp, tgt_mask)
                        dec_inp = torch.cat((dec_inp, pred[:, -1:, :]), 1)

                    pred_denorm = denormalisation(pred, translation, amplitude)
                    target_denorm = denormalisation(target, translation, amplitude)

                    diff_x = pred_denorm[:, :, 0] - target_denorm[:, :, 0]  # B x 5 x 1
                    distances_x = torch.abs(diff_x)
                    loss_x = distances_x.mean(dim=0)  # batch average

                    diff_y = pred_denorm[:, :, 1] - target_denorm[:, :, 1]  # B x 5 x 1
                    distances_y = torch.abs(diff_y)
                    loss_y = distances_y.mean(dim=0)  # batch average

                    diff_z = pred_denorm[:, :, 2] - target_denorm[:, :, 2]  # B x 5 x 1
                    distances_z = torch.abs(diff_z)
                    loss_z = distances_z.mean(dim=0)  # batch average

                    epoch_loss_x.append(loss_x.cpu().tolist())
                    epoch_loss_y.append(loss_y.cpu().tolist())
                    epoch_loss_z.append(loss_z.cpu().tolist())
                    nbr_batch += 1

                pbar.update(1)

            # 70 x 5
            mean_x = np.mean(np.array(epoch_loss_x), axis=0)
            mean_y = np.mean(np.array(epoch_loss_y), axis=0)
            mean_z = np.mean(np.array(epoch_loss_z), axis=0)

            ade_x = np.mean(mean_x, axis=0)
            ade_y = np.mean(mean_y, axis=0)
            ade_z = np.mean(mean_z, axis=0)

            ade = np.sqrt(ade_x ** 2 + ade_y ** 2 + ade_z ** 2)
            fde = np.sqrt(mean_x[-1] ** 2 + mean_y[-1] ** 2 + mean_z[-1] ** 2)

            dico_intra = {50: 2, 100: 62, 200: 122, 300: 182, 400: 242, 500: 302, 600: 362, 700: 422, 800: 482,
                          900: 542, 1000: 602, 1500: 2}
            col = dico_intra[50] + 6 * int((cfg.seed - 1))

            wb = load_workbook(saving_path)

            if cfg.patient < 16:
                sheet_x = wb[f'FDG{list(test_part)[-1]}1_X']
                sheet_y = wb[f'FDG{list(test_part)[-1]}1_Y']
                sheet_z = wb[f'FDG{list(test_part)[-1]}1_Z']
            else:
                sheet_x = wb[f'NO_CPAP{list(test_part)[-1]}1_X']
                sheet_y = wb[f'NO_CPAP{list(test_part)[-1]}1_Y']
                sheet_z = wb[f'NO_CPAP{list(test_part)[-1]}1_Z']

            sheet_x.cell(row=patient + 1, column=1).value = f'Patient_{patient}'
            sheet_y.cell(row=patient + 1, column=1).value = f'Patient_{patient}'
            sheet_z.cell(row=patient + 1, column=1).value = f'Patient_{patient}'
            for i in range(cfg.training.params.horizon):
                sheet_x.cell(row=patient + 1, column=col + i).value = mean_x[i]
                sheet_y.cell(row=patient + 1, column=col + i).value = mean_y[i]
                sheet_z.cell(row=patient + 1, column=col + i).value = mean_z[i]

            wb.save(saving_path)

    if verbose:
        print(f"X: {ade_x}, Y: {ade_y}, Z: {ade_z}, AVG: {ade}")


@hydra.main(version_base="1.3", config_path="conf", config_name="config")
def main(cfg):
    seed = cfg.seed
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)

    if cfg.multi is False:
        data_path = f'/Benson_DATA2/grotsart/Medical_physics/PS-models/{cfg.data_quantity}/{cfg.training.params.width}-' \
                    f'{cfg.training.params.patch_size}/'
    else:
        data_path = f'/Benson_DATA2/grotsart/Medical_physics/MP-models/{cfg.data_quantity}/{cfg.training.params.width}-' \
                    f'{cfg.training.params.patch_size}/'
    if cfg.training.params.strategy == 'T1':
        model_path = f'Patient_{cfg.patient}-LR-{cfg.training.params.learning_rate}-epoch-{cfg.inference.params.epoch}-' \
                     f'{cfg.seed}.pt'
    elif cfg.training.params.strategy == 'T2':
        model_path = f'T2-Patient_{cfg.patient}-LR-{cfg.training.params.learning_rate}-epoch-{cfg.inference.params.epoch}-' \
                     f'{cfg.seed}.pt'
    else:
        raise NotImplementedError

    net = Net(cfg=cfg)
    checkpoint = torch.load(data_path + model_path, map_location=get_default_device(cfg.device))
    net.load_state_dict(checkpoint['model_state_dict'])
    net = to_device(net, cfg.device)  # utile ?

    inference_random(model=net, cfg=cfg, test_part='T1', verbose=True)
    inference_random(model=net, cfg=cfg, test_part='T2', verbose=True)


if __name__ == "__main__":
    main()
